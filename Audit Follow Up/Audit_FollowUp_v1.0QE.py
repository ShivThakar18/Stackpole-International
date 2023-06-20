
# *------------------------------------ Shiv Thakar - Quality Engineering Intern ------------------------------------- #
# *--------------------------------------------- Stackpole International --------------------------------------------- #
# *------------------------------------ Daily Quality Audit Follow Up Application ------------------------------------ #
# *---------------------------------------------- Date: April 4th, 2023 ---------------------------------------------- #
# !------------------------------------------------------------------------------------------------------------------- #
# !                                               Import Python Modules                                                #
# !------------------------------------------------------------------------------------------------------------------- #
from os import path, rename,stat
from glob import glob
from tkinter import *
import tkinter.messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from shutil import copy, move
from time import sleep
from openpyxl.styles import Alignment
import copy
from os import remove
from datetime import datetime
# !------------------------------------------------------------------------------------------------------------------- #
# !                                                 Define Variables                                                   #
# !------------------------------------------------------------------------------------------------------------------- #
LITMUS_DRIVE = "L:\\ShivDataOutput\\Daily_Quality_Audits\\Follow_Up\\"                      # file source
DATA_FOLDER = "N:\\Quality\\Daily Quality Audits\\"  # file dest
FILE_LOC = DATA_FOLDER + ""
XLSX_FILE = "Audit Follow Up List.xlsx"
# !------------------------------------------------------------------------------------------------------------------- #
# !                                                     Functions                                                      #
# !------------------------------------------------------------------------------------------------------------------- #                                        
# ?------------------------------------------------ Parse and Extract ------------------------------------------------ #
def parseFile(FILE):

    m_file = open(FILE,'r')                                       # open FollowUp.txt file for reading
    m_data = m_file.read().splitlines()                           # read file
    m_file.close()                                                # close file

    if(stat(FILE).st_size == 0):                                  # check if file is empty
        print("# --- File Empty")
        return                                                    # return and end the program

    for i in range(len(m_data)):                                  # iterate through each line
        m_data[i] = m_data[i][1:-1]                               # remove {} from string
    
    data = []
    temp = []

    for m in m_data:                                              # iterate through m_data
        temp = m.replace(",\"","|\"")                             # replace all ," with |"
        temp = temp.split("|")                                    # split using |
        print(temp)
        actual = []                                 
        x = 0                                                       

        if("Specific Requests" in temp[1] or "Critical Calls" in temp[1]):  # if the section is SR or CC
            date = temp[3].split(":")[1]
            date_time = datetime.fromtimestamp(date/1000)
            date_str = date_time.strftime("%d-%m-%Y")

            actual.append(date_str)                             # date

            shift = temp[2].split(":")[1]
            if(temp == '1'):
                shift = 'Day'
            elif(temp == '2'):
                shift = 'Afternoon'
            elif(temp == '3'):
                shift = 'Night'
            actual.append(shift)                                # Shift

            actual.append(temp[0].split(":")[1][1:-1])          # Report Number
            actual.append(temp[1].split(":")[1][1:-1])          # Section
            actual.append(temp[11].split(":")[1][1:-1])         # Person
            actual.append(temp[4].split(":")[1][1:-1])          # Quality Tech
            actual.append(temp[5].split(":")[1][1:-1])          # Description
            actual.append(temp[8].split(":")[1][1:-1])          # Action Comments
            actual.append(temp[10].split(":")[1][1:-1])         # Comments

        elif("Bypass List" in temp[1]):                         # if the section is BL or Other NDT Cells
            
            date = temp[3].split(":")[1]
            date_time = datetime.fromtimestamp(float(date)/1000)
            date_str = date_time.strftime("%d-%m-%Y")

            actual.append(date_str)                             # date

            shift = temp[2].split(":")[1]

            if(shift == '1'):
                shift = 'Day'
            elif(shift == '2'):
                shift = 'Afternoon'
            elif(shift == '3'):
                shift = 'Night'
            
            actual.append(shift)                                # Shift

            actual.append(temp[0].split(":")[1][1:-1])          # Report Number
            actual.append(temp[1].split(":")[1][1:-1])          # Section
            
            process = temp[6].split(":")[1][1:-1]               # extract process

            for t in temp:                                      # iterate through all the elements
                if(':' not in t):                               # find rows without ":"
                    x = x + 1                                   # increment index variable x
                    process = process + t                       # concatenate string
            
            actual.append(temp[11 + x].split(":")[1][1:-1])         # Person
            actual.append(temp[4].split(":")[1][1:-1])              # Quality Tech

            part = temp[8 + x].split(":")[1][1:-1]                  # extract part

            if("\"" in temp[6].split(":")[1][1:-1]):                    # if " is in the process string
                process_format = process.replace("\",\""," \\ ")        # replace "," with \
                process_format = process_format.replace("\"\"",", ")    # replace "" with space
                process_format = process_format.replace("\""," ")       # replace " with space
                process_format = process_format.replace("]","")         # replace ] with blank
                process_format = process_format.replace(" ", "",1)      # replace the first space with blank
                actual.append(process_format + " - " + part)        # Description
            actual.append(temp[9 + x].split(":")[1][1:-1])          # Action
            actual.append(temp[10 + x].split(":")[1][1:-1])         # Comments

        data.append(actual)                                         # append to list
        
    write_excel(data)                                               # call write excel
    """ txt = open(FILE_LOC + "FollowUp.txt",'w')                       # rewrite as blank file
    txt.close()  """
# ?------------------------------------------------ Write Excel File ------------------------------------------------- #
def write_excel(data):
    global DATA_FOLDER, XLSX_FILE
    print("# ----------------------------------------- Writing Excel ---------------------------------------- #")
    while(True):
        try:
            read_template = pd.read_excel(DATA_FOLDER + XLSX_FILE, engine='openpyxl') # open the template as pandas dataframe
            wb = load_workbook(DATA_FOLDER + XLSX_FILE)                               # load the workbook
            ws = wb.active                                                            # get the last active row
            next_row = ws.max_row + 1                                                 # get the next blank row after active row
            break                                                                     # break to rest of program once workbook is available
        except PermissionError:
            print("# --- Permission Error: Waiting for file to be accessible")
            continue                                                                  # keep trying to open workbook until available

    for d in data:      # iterate through data list (list of lists)
        read_template.at[next_row,'Unnamed: 0'] = d[0]                          # Date
        read_template.at[next_row,'Unnamed: 1'] = d[1]                          # Shift
        read_template.at[next_row,'Unnamed: 2'] = d[2]                          # Report Num
        read_template.at[next_row,'Unnamed: 3'] = d[3]                          # Section
        read_template.at[next_row,'Unnamed: 4'] = d[4]                          # Follow Up Person
        read_template.at[next_row,'Daily Quality Audit Follow Up List'] = d[5]  # Quality Tech
        read_template.at[next_row,'Unnamed: 6'] = d[6]                          # Description
        read_template.at[next_row,'Unnamed: 7'] = d[7]                          # Action Taken
        read_template.at[next_row,'http://pmdadashboard/d/DluY5Ph4z/daily-shift-quality-audit-report?orgId=1&var-QA_ReportNo=All&from=now%2Fy&to=now%2Fy'] = d[8]
                                                                                # Follow Up Comments
        read_template.at[next_row,'Unnamed: 9'] = 'N'                           # Completed? 
        
        next_row = next_row + 1                                                 # increment row, to write on next row
    
    df_columns = read_template.shape[1]                                         # get columns                                                                               

    for r, row in enumerate(dataframe_to_rows(read_template,index=False, header=False),2):     # write to each points from dataframe to cells in excel
        for c in range(0,df_columns):
            try:
                ws.cell(row = r, column= c + 1).value = row[c]
            except AttributeError:
                pass
    
    for row in ws.iter_rows():         # apply formatting to each cell                                     
        for cell in row:
            cell.alignment = cell.alignment.copy(wrapText=True)         # apply text wrapping
            cell.alignment = cell.alignment.copy(vertical='center')     # apply vertical center

    while(True):
        try:
            wb.save(DATA_FOLDER + XLSX_FILE)                            # save the file
            break
        except PermissionError:                                         # keep trying until saving is possible
            continue
# !------------------------------------------------------------------------------------------------------------------- #
# !                                                   Function Call                                                    #
# !------------------------------------------------------------------------------------------------------------------- #
while(True):
    try:    
        file = glob(FILE_LOC + "*.txt")[0]                         # search for file
        print("# ------------------------------------------ File Found ------------------------------------------ #")         
        print("# --- Parsing File")
        parseFile(file)                                            # call parsefile function
        print("# --- Update Complete")
        break                                                      # if excel file was done and saved
    except:                     
        print("# ---" + str(Exception))
        sleep(5)
        pass 